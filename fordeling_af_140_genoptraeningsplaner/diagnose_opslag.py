"""Diagnosis code lookup from Excel spreadsheet."""

from openpyxl import load_workbook


def indlæs_diagnosekoder(excel_path: str) -> list[dict]:
    """Load diagnosis search terms from Excel and sort by search text length descending.

    Reads the "Diagnoser" sheet, transposes it (first row = field names,
    subsequent columns = values), filters out empty entries, and returns
    sorted by search text length so longest matches are tried first.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Diagnoser"]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    kategorier = []

    for kategori in rows[0]:
        if kategori:
            kategorier.append(str(kategori).strip())

    diagnosekoder = []

    for row in rows[1:]:
        for idx, kategori in enumerate(kategorier):
            if idx < len(row):
                value = row[idx]
                if value:
                    diagnosekoder.append(
                        {
                            "Diagnose": kategori,
                            "Sogetekst": str(value).strip().lower(),
                        }
                    )

    diagnosekoder.sort(key=lambda d: len(d["Sogetekst"]), reverse=True)
    return diagnosekoder


def find_diagnose_kategori(
    diagnoser: list[dict], diagnosekoder: list[dict]
) -> tuple[str, str]:
    """Match GOP diagnoses against the lookup table.

    For short search terms (<=3 chars), does exact word matching against
    all words in diagnosis texts/codes. For longer terms, does substring
    matching against the raw diagnosis text and code.

    Returns (category, basal_avanceret_diagnose).
    """
    # Build word list from all diagnoses for short-term matching
    alle_ord = []
    for diag in diagnoser:
        text = f"{diag.get('Tekst', '')} {diag.get('Kode', '')} {diag.get('Type', '')}"
        alle_ord.extend(text.lower().split())

    # Try each search term (longest first)
    kategori = None
    for dk in diagnosekoder:
        sogetekst = dk["Sogetekst"]

        if len(sogetekst) <= 3:
            # Short search term: exact word match
            if sogetekst in alle_ord:
                kategori = dk["Diagnose"]
                break
        else:
            # Long search term: substring match in any diagnosis
            for diag in diagnoser:
                kode = diag.get("Kode", "").lower()
                tekst = diag.get("Tekst", "").lower()
                if sogetekst in kode or sogetekst in tekst:
                    kategori = dk["Diagnose"]
                    break
            if kategori:
                break

    if kategori is None:
        kategori = "Andet"

    # Determine basal/avanceret diagnosis for Sue model
    basal_avanceret = _bestem_basal_avanceret_diagnose(kategori, diagnoser)

    return kategori, basal_avanceret


def _bestem_basal_avanceret_diagnose(kategori: str, diagnoser: list[dict]) -> str:
    """Determine the basal/avanceret diagnosis label for the Sue treatment model."""
    if kategori in ("Cancer", "Amputation", "Neurologi"):
        return kategori

    # Check for lænderygbesværer
    for diag in diagnoser:
        combined = (
            f"{diag.get('Type', '')} {diag.get('Tekst', '')} {diag.get('Kode', '')}"
        )
        if (
            "lænderygbesværer" in combined.lower()
            or "laenderygbesvaer" in combined.lower()
        ):
            return "Kronisk lænderygbesværer"

    return "Ukendt"
